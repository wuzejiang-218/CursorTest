#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate XMind test cases and smoke Excel from structured JSON."""

from __future__ import annotations

import json
import math
import sys
import uuid
import zipfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: pip install openpyxl") from exc


SMOKE_COLUMNS = [
    "案例需求",
    "用例标题",
    "测试数据",
    "前置条件",
    "测试步骤",
    "预期结果",
    "执行人员",
    "冒烟结果",
    "不通过原因",
    "开发",
]

PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}


def _id() -> str:
    return str(uuid.uuid4()).replace("-", "")[:24]


def normalize_lines(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    return str(value)


def safe_filename(value: str) -> str:
    return "".join(ch for ch in value if ch not in r'\/:*?"<>|').strip() or "测试用例"


def get_test_data(case: dict[str, Any]) -> str:
    return normalize_lines(
        case.get("test_data")
        or case.get("testData")
        or case.get("测试数据")
        or case.get("data")
        or case.get("precondition")
    )


def get_case_id(case: dict[str, Any], index: int, prefix: str) -> str:
    explicit = (
        case.get("case_id")
        or case.get("caseId")
        or case.get("id")
        or case.get("用例ID")
        or case.get("案例编码")
    )
    if explicit:
        return str(explicit)
    return f"{prefix}-{index:03d}"


def format_case_title(case: dict[str, Any], index: int, prefix: str) -> str:
    return f"{get_case_id(case, index, prefix)} {case.get('title', '未命名用例')}"


def topic(title: str, children: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    node: dict[str, Any] = {"id": _id(), "title": title}
    if children:
        node["children"] = {"attached": children}
    return node


def mark_smoke_cases(cases: list[dict[str, Any]], smoke_ratio: float) -> None:
    if not cases:
        return

    explicit_count = sum(1 for case in cases if case.get("is_smoke") is True)
    if explicit_count:
        return

    target = max(5, math.ceil(len(cases) * smoke_ratio))
    target = min(len(cases), target)

    sorted_cases = sorted(
        cases,
        key=lambda case: (
            PRIORITY_ORDER.get(str(case.get("priority", "P3")).upper(), 99),
            str(case.get("module", "")),
            str(case.get("title", "")),
        ),
    )
    for index, case in enumerate(sorted_cases):
        case["is_smoke"] = index < target


def build_xmind_content(
    title: str,
    cases: list[dict[str, Any]],
    case_id_prefix: str,
) -> list[dict[str, Any]]:
    modules: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    indexed_cases = list(enumerate(cases, 1))
    for case_index, case in indexed_cases:
        modules.setdefault(str(case.get("module") or "未分组"), []).append((case_index, case))

    module_topics: list[dict[str, Any]] = []
    for module_name, module_cases in modules.items():
        scenario_topics = []
        for case_index, case in module_cases:
            smoke_prefix = "[SMOKE] " if case.get("is_smoke") else ""
            scenario_title = f"{smoke_prefix}{format_case_title(case, case_index, case_id_prefix)}"
            scenario_topics.append(
                topic(
                    scenario_title,
                    [
                        topic(f"测试数据：{get_test_data(case) or '无'}"),
                        topic(f"前置：{normalize_lines(case.get('precondition')) or '无'}"),
                        topic(f"步骤：{normalize_lines(case.get('steps')) or '待补充'}"),
                        topic(f"预期：{normalize_lines(case.get('expected')) or '待补充'}"),
                        topic(f"优先级：{case.get('priority', 'P2')}"),
                    ],
                )
            )
        module_topics.append(topic(module_name, scenario_topics))

    sheet_id = _id()
    return [
        {
            "id": sheet_id,
            "class": "sheet",
            "title": title,
            "rootTopic": {
                "id": _id(),
                "class": "topic",
                "title": title,
                "structureClass": "org.xmind.ui.map.unbalanced",
                "children": {"attached": module_topics},
            },
        }
    ]


def write_xmind(path: Path, content: list[dict[str, Any]]) -> None:
    metadata = {
        "creator": {"name": "testcase-xmind-smoke-output", "version": "1.0"},
        "activeSheetId": content[0]["id"],
    }
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))


def write_smoke_excel(path: Path, cases: list[dict[str, Any]], case_id_prefix: str) -> None:
    smoke_cases = [case for case in cases if case.get("is_smoke")]
    wb = Workbook()
    ws = wb.active
    ws.title = "冒烟案例"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    smoke_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

    for col_index, column in enumerate(SMOKE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, case in enumerate(smoke_cases, 2):
        row = [
            case.get("requirement") or case.get("module") or "",
            case.get("title") or "",
            get_test_data(case),
            normalize_lines(case.get("precondition")),
            normalize_lines(case.get("steps")),
            normalize_lines(case.get("expected")),
            case.get("executor", ""),
            case.get("smoke_result", ""),
            case.get("failed_reason", ""),
            case.get("developer", ""),
        ]
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.fill = smoke_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [22, 42, 34, 28, 48, 48, 14, 14, 28, 16]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python generate_xmind_smoke_excel.py cases.json [output_dir]")
        return 1

    input_path = Path(sys.argv[1]).resolve()
    output_dir = Path(sys.argv[2]).resolve() if len(sys.argv) >= 3 else input_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    payload = json.loads(input_path.read_text(encoding="utf-8"))
    cases = payload.get("cases", [])
    if not isinstance(cases, list) or not cases:
        raise SystemExit("cases.json must contain a non-empty 'cases' array")

    smoke_ratio = float(payload.get("smoke_ratio", 0.2))
    mark_smoke_cases(cases, smoke_ratio)

    title = str(payload.get("title") or "测试用例")
    output_prefix = safe_filename(str(payload.get("output_prefix") or title))
    case_id_prefix = str(payload.get("case_id_prefix") or "TC")

    xmind_path = output_dir / f"{output_prefix}-测试用例.xmind"
    excel_path = output_dir / f"{output_prefix}-冒烟测试用例.xlsx"

    write_xmind(xmind_path, build_xmind_content(title, cases, case_id_prefix))
    write_smoke_excel(excel_path, cases, case_id_prefix)

    smoke_count = sum(1 for case in cases if case.get("is_smoke"))
    print(f"Generated: {xmind_path}")
    print(f"Generated: {excel_path}")
    print(f"Smoke cases: {smoke_count}/{len(cases)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
